# coding: latin-1
# -*- coding: utf-8 -*-
import pyautogui
import pyperclip
import time
import logging
import telebot
import os
import numpy as np
from config import *
import openpyxl
import win32com.client
import warnings
warnings.simplefilter("ignore", UserWarning)
from psycopg2 import sql, extras
import glob
#import psycopg2
import pandas

tupla_inicar_sesion = (
            'cancelar_chrome.png',
            'eliminar_sesion.png',
            'email.png',
            'abandonar.png',
            'email.png',
            )

tupla_filtro = (
            'logo_censel.png',
            'logo_reportes_web.png',
            'eventos.png', 
            'reporte_eventos_por_fecha.png',
            'filtros.png',) 

tupla_postformulario = (    
            'buscar.png',
            'exportar.png',
            'exportar_a_csv.png',)

horarios_procesos = [
    #nombre_proceso                 hora_ejecucion  hora_desde hora_hasta          codigo_alarma                    columnas_excel1    columnas_excel2      columnas_excel3         columnas_excel4    columnas_excel5     tabla                         
    ('baterias1',  fecha_desde, fecha_hasta,'8:05 PM', '00:00', '12:00', 'FALLO DE BATERIA / BATTERY FAILURE - LOW', 'cue_ncuenta', '',                  '',                    '',                    '',        'replica_registro_codigos_seguridad'),
    ('baterias2',  fecha_desde, fecha_hasta,'12:00 PM','12:00', '00:00', 'FALLO DE BATERIA / BATTERY FAILURE - LOW', 'cue_ncuenta', '',                  '',                    '',                    '',        'replica_registro_codigos_seguridad'),
    ('intrusion',  fecha_ayer,  fecha_hoy,  '8:05 AM', '19:00', '07:00', 'INTRUSION - BUR',                          'cue_ncuenta', 'rec_czona',         'rec_tFechaProceso',   'rec_tFechaRecepcion', '_puerto', 'replica_seg_control_novedades'),
    ('fallo_test', fecha_ayer,  fecha_hoy,  '6:43 PM', '19:00', '07:00', 'FALLO DE TEST / TEST FAIL - FTS',          'cue_ncuenta', 'rec_tFechaProceso', 'rec_tFechaRecepcion', 'tablaDatos',          '',        'replica_seg_control_novedades'),
    ('panico',     fecha_ayer,  fecha_ayer, '8:15 AM', '00:00', '23:50', 'PANICO - PAN',                             'cue_ncuenta', 'rec_czona',         'rec_tFechaProceso',   'rec_tFechaRecepcion', '_puerto', 'replica_seg_control_novedades') #funciona
    ] #hora_actual
# count: nos devuelve el numero de veces que se repite un elemento
# index: Nos devuelve la posicion de la primera aparicion de un elemento.c 

# TELEGRAM BOT
def mensaje_telegram(mensaje, ruta_imagen, nombre_imagen, carpeta, nombre_proceso, horario, img):
    TELEGRAM_BOT_TOKEN = '6232135002:AAGPl356BEAbpzSQlgomBQi45YBUZJk136Q'
    TELEGRAM_CHAT_ID = '7411433556'
    contenido_mensaje = hora_actual
    bot = telebot.TeleBot(TELEGRAM_BOT_TOKEN)
    if mensaje == 'inicio':
        contenido_mensaje += (f"\n!! SE INICIA PROCESO DE CENSEL A LAS {hora_actual} !!")
        bot.send_message(chat_id=TELEGRAM_CHAT_ID, text=contenido_mensaje)       
    elif mensaje == 'ingreso_censel':
        contenido_mensaje += (f'\nSe ingresó al sistema de Censel correctamente')
        with open(ruta_imagen, 'rb') as img_file:
            img_data = img_file.read()
            bot.send_photo(chat_id=TELEGRAM_CHAT_ID, photo=img_data, caption=contenido_mensaje)
    elif mensaje == 'reportes_web':
        contenido_mensaje += (f'\nSe ingresó a <b>Reportes Web</b> correctamente')
        with open(ruta_imagen, 'rb') as img_file:
            img_data = img_file.read()
            bot.send_photo(chat_id=TELEGRAM_CHAT_ID, photo=img_data, caption=contenido_mensaje, parse_mode='HTML')
    elif mensaje == 'error_ruta_imagen':
        contenido_mensaje += (f'\nNo existe la imagen ({nombre_imagen}) en la carpeta: ({carpeta})')
        bot.send_message(chat_id=TELEGRAM_CHAT_ID, text=contenido_mensaje)
    elif mensaje == 'filtrando_proceso':
        contenido_mensaje += (f'\n Se filtra el proceso de {nombre_proceso} al horario: {horario}')
        with open(ruta_imagen, 'rb') as img_file:
            img_data = img_file.read()
        bot.send_photo(chat_id=TELEGRAM_CHAT_ID, photo=img_data, caption=contenido_mensaje, parse_mode='HTML')
    elif mensaje == img:
        contenido_mensaje += (f'\nSe ingresó a <b>{mensaje}</b> correctamente')
        with open(ruta_imagen, 'rb') as img_file:
            img_data = img_file.read()
        bot.send_photo(chat_id=TELEGRAM_CHAT_ID, photo=img_data, caption=contenido_mensaje, parse_mode='HTML')
    elif mensaje == 'no hay eventos':
        contenido_mensaje += (f'\n No se encontraron registros al momento de hacer el filtro en el proceso de <b>{nombre_proceso}</b>  \nSe cancela la ejecucion del proceso')
        with open(ruta_imagen, 'rb') as img_file:
            img_data = img_file.read()
        bot.send_photo(chat_id=TELEGRAM_CHAT_ID, photo=img_data, caption=contenido_mensaje, parse_mode='HTML')
    elif mensaje == 'horarios_diferentes':
        contenido_mensaje += (f'\n La hora de ejecucion: {horario} del proceso de {nombre_proceso} no coincide con la hora actual, favor validar.')
        bot.send_message(chat_id=TELEGRAM_CHAT_ID, text=contenido_mensaje)
    elif mensaje == 'cerrar_chrome':
        contenido_mensaje += (f'\n Se finaliza el proceso en la pagina de censel y se cierra el navegador. \n\n Se inicia con el procesamiento de la iformacion en el archivo excel')
        bot.send_message(chat_id=TELEGRAM_CHAT_ID, text=contenido_mensaje)
    elif mensaje == 'sesion_encontrada':
        contenido_mensaje += (f'Se encontro una sesion abierta en el aplicativo de censel. \nSe cierra sesion y se inicia nuevamente.')
        with open(ruta_imagen, 'rb') as img_file:
            img_data = img_file.read()
        bot.send_photo(chat_id=TELEGRAM_CHAT_ID, photo=img_data, caption=contenido_mensaje, parse_mode='HTML')

def obtener_coordenadas_imagen_pantalla(ruta_imagen):
    try:
        coordenadas = pyautogui.locateOnScreen(ruta_imagen, confidence=0.9)
        #if coordenadas:
            #coordenadas = (coordenadas.left, coordenadas.top)
        logging.info(f"Ruta imagen {ruta_imagen}. Coordenadas: {coordenadas}")
        return coordenadas
    except Exception:
        logging.error(f'No se encontro la imagen {ruta_imagen} en la pantalla')
        logging.error(Exception)
        logging.error('Error en la funcion: obtener_coordenadas_imagen_pantalla.')
        return "error"

def obtener_ruta_imagenes(nombre_imagen):    
    try:                                     
        if "pantallazo" in nombre_imagen:
            carpeta = "screenshots"
        else:   
            carpeta = "img"

        ruta_imagen = os.path.join(os.getcwd(), carpeta, nombre_imagen)
        
        if os.path.exists(ruta_imagen) and os.path.isfile(ruta_imagen): 
            return ruta_imagen
        logging.error(f"No existe la imagen ({nombre_imagen}) en la carpeta ({carpeta})")
        return mensaje_telegram('error_ruta_imagen', ruta_imagen, nombre_imagen, carpeta, None, None, None)
    except Exception as e:
        logging.error(e)
        logging.error('Error en la funcion: obtener_ruta_imagenes.')

def obtener_captura_pantalla(nombre_captura, carpeta):
    ruta_actual = os.getcwd()
    ruta_errores = os.path.join(ruta_actual, carpeta)
    nombre_base, extension = os.path.splitext(nombre_captura)
    nombre_modificado = f'{nombre_base}_pantallazo{extension}'
    
    ruta_captura = os.path.join(ruta_errores, nombre_modificado)
    try:
        pyautogui.screenshot(ruta_captura)
        logging.info(f"Captura de pantalla exitosa --> Nombre: {nombre_captura} --> carpeta: {carpeta}")
        return ruta_captura
    except Exception as e:
        logging.error(f" Error al capturar: {e}")
        logging.error(f'Error con la ruta de la captura: {ruta_captura}')
        logging.error('Error en la funcion: obtener_captura_pantalla')

def iniciar_filtro(img):
    try:
        error = '_error'
        if not img == 'eventos.png':
            ruta_imagen = obtener_ruta_imagenes(img)
            coordenadas_ruta_imagen = obtener_coordenadas_imagen_pantalla(ruta_imagen)
            if coordenadas_ruta_imagen:
                time.sleep(2)
                pyautogui.click(coordenadas_ruta_imagen)
                time.sleep(2)
                logging.info(f'Se hizo click en: {img}')
                ruta_captura = obtener_captura_pantalla(img, 'screenshots')
                ruta_imagen = obtener_ruta_imagenes(ruta_captura)
                if ruta_imagen:
                    mensaje_telegram(img, ruta_imagen, None, None, None, None, img)
                    return True
                else:
                    img = img + error
                    mensaje_telegram(img, ruta_imagen, None, None, None, None, img)
                    return False
        elif img == 'abandonar.png':
            ruta_imagen = obtener_ruta_imagenes(img)
            coordenadas_ruta_imagen = obtener_coordenadas_imagen_pantalla(ruta_imagen)
            if coordenadas_ruta_imagen:
                time.sleep(2)
                logging.info(f'Se hizo click en: {img}')
                ruta_captura = obtener_captura_pantalla(img, 'screenshots')
                ruta_imagen = obtener_ruta_imagenes(ruta_captura)
                if ruta_imagen:
                    mensaje_telegram(img, ruta_imagen, None, None, None, None, img)
                    iniciar_sesion()
                    return True
                else:
                    img = img + error
                    mensaje_telegram(img, ruta_imagen, None, None, None, None, img)
                    return False
        else:
            ruta_imagen = obtener_ruta_imagenes(img)
            coordenadas_ruta_imagen = obtener_coordenadas_imagen_pantalla(ruta_imagen)
            if coordenadas_ruta_imagen:
                time.sleep(2)
                pyautogui.doubleClick(coordenadas_ruta_imagen)
                time.sleep(2)
                logging.info(f'Se hizo click en: {img}')
                ruta_captura = obtener_captura_pantalla(img, 'screenshots')
                ruta_imagen = obtener_ruta_imagenes(ruta_captura)
                if ruta_imagen:
                    mensaje_telegram(img, ruta_imagen, None, None, None, None, img)
                    return True
                else:
                    img = img + error
                    mensaje_telegram(img, ruta_imagen, None, None, None, None, img)
                    return False
    except Exception as e:
        logging.error(f'Ocurrio un error al tratar de encontrar la imagen en la pantalla y dar click: {e}')
        logging.error('Ocurrio un error en la funcion: iniciar_filtro.')
        logging.error(f'Parametro recibido en iniciar_filtro: {img}')

# def consultarlistaAbonados(valores, tabla, valores_rec_czona, valores_rec_tFechaProceso, valores_rec_tFechaRecepcion, valores_puerto, nombre_proceso):
#     try:
#         print('dentro de la funcion: consultarlistaAbonados')
#         contador = 0
#         print('Contador ->',contador)
#         cur = conexion.cursor()
        
#         # Insertar los valores de col_values directamente
#         if(tabla == 'replica_registro_codigos_seguridad'):
#             for valor in valores:
#                 # Construir la consulta de manera segura
#                 query = sql.SQL(f"INSERT INTO {tabla} (codigo_seguridad, fecha, tipo) VALUES (%s, %s, %s) RETURNING id_codigo").format(tabla=sql.Identifier(tabla)
#                 )

#                 # Ejecutar la consulta pasando los valores como parámetros
#                 cur.execute(query, (valor, fecha_ayer, '2'))

#                 # Obtener el ID insertado
                
#                 id_insertado = cur.fetchone()[0]  # Obtener el ID insertado
#                 if id_insertado:
#                     contador+=1
#                 else:
#                     print("Inserción fallida, no se retornó un ID ", valor)
#         else:
#             for valor in zip(valores, valores_rec_czona, valores_rec_tFechaProceso, valores_rec_tFechaRecepcion, valores_puerto,):
#                 cue_ncuenta, rec_czona, rec_tFechaProceso, rec_tFechaRecepcion, puerto, = valor
                
#                 query = sql.SQL(f"INSERT INTO {tabla} (nombre_novedad,tipo_novedad,tipo_sensor,puerto_nov,fecha_proceso,fecha_recepcion,codigo_abonado,estado_gestion,fecha_novedad) VALUES (%s, %s, %s,%s, %s, %s,%s, %s, %s) RETURNING id_nov").format(tabla=sql.Identifier(tabla))

#                 # Ejecutar la consulta pasando los valores como parámetros
#                 cur.execute(query, (nombre_proceso,'1',rec_czona,puerto,rec_tFechaProceso,rec_tFechaRecepcion,cue_ncuenta,'1',fecha_ayer))

#                 # Obtener el ID insertado
                
#                 id_insertado = cur.fetchone()[0]  # Obtener el ID insertado
#                 if id_insertado:
#                     contador+=1
#                 else:
#                     print("Inserción fallida, no se retornó un ID ", valor)
                
#                 sql_insert = f"""
#                 INSERT INTO replica_seg_control_novedades (nombre_novedad,tipo_novedad,tipo_sensor,puerto_nov,fecha_proceso,fecha_recepcion,codigo_abonado,estado_gestion,fecha_novedad)
#                 VALUES ('INTRUSION','1','{rec_czona}','{puerto}','{rec_tFechaProceso}','{rec_tFechaRecepcion}','{cue_ncuenta}','1','{fecha_ayer}');
#                 """
#                 logging.info(f"Ejecutando SQL: {sql_insert}")
#                 cur.execute(sql_insert,(fecha_ayer))
#                 conexion.commit()
#         cur.close()
#         conexion.close()

    #     # Enviar mensaje de exito a Telegram
    #     print(f"fue exitosa la insertacion en la base de datos, Número de filas insertada {contador}")
    # except Exception as e:
    #     logging.error("Error al insertar datos:", e)
    #     # Enviar mensaje de error a Telegram
    #     print(f"Error al insertar datos: {e}")
    # finally:
    #     if os.path.exists(r'C:\Users\auxsenadesarrollo\Downloads\reportehistoricohtml.xlsx'):
    #         os.remove(r'C:\Users\auxsenadesarrollo\Downloads\reportehistoricohtml.xlsx')
    #         logging.info(r"Archivo C:\Users\auxsenadesarrollo\Downloads\reportehistoricohtml.xlsx eliminado despues de 1 minuto.")
    #         # enviar_mensaje_telegram(f"Archivo C:\Users\auxsenadesarrollo\Downloads\reportehistoricohtml.xlsx eliminado despuÃ©s de 1 minuto.")
    #     else:
    #         logging.warning(r"El archivo C:\Users\auxsenadesarrollo\Downloads\reportehistoricohtml.xlsx no existe.")

# def procesar_archivo_excel(ruta_archivo, tabla, nombre_proceso):
    # try:
    #     print('dentro de la funcion: procesar_archivo_excel')
    #     # Abrir el archivo de Excel
    #     workbook = openpyxl.load_workbook(ruta_archivo)
    #     sheet = workbook.active

    #     # Buscar la columna que contiene "cue_ncuenta" 
    #     def obtener_valores_columna(nombre_columna):
    #         target_col = None
    #         for cell in sheet[1]:  
    #             if cell.value == nombre_columna:
    #                 target_col = cell.column
    #                 break
    #         if not target_col:
    #             print(f"No se encontró la columna con '{nombre_columna}'.")
    #             return []
    #         else:
    #             col_values = [
    #                 sheet.cell(row=row, column=target_col).value 
    #                 for row in range(2, sheet.max_row + 1) 
    #                 if sheet.cell(row=row, column=target_col).value is not None
    #                 ]
    #             col_values = list(dict.fromkeys(col_values)) 
    #             return col_values
            
    #     def limpiar_espacios(valores):
    #         return [valor.strip() if isinstance(valor, str) else valor for valor in valores]

    #     col_cue_ncuenta = limpiar_espacios(obtener_valores_columna("cue_ncuenta"))
    #     col_rec_czona = limpiar_espacios(obtener_valores_columna("rec_czona"))
    #     col_rec_tFechaProceso = limpiar_espacios(obtener_valores_columna("rec_tFechaProceso"))
    #     col_rec_tFechaRecepcion = limpiar_espacios(obtener_valores_columna("rec_tFechaRecepcion"))
    #     col_puerto = limpiar_espacios(obtener_valores_columna("_puerto"))
    #     # Imprimir los valores guardados
    #     logging.info("Valores únicos de la columna 'cue_ncuenta':")

    #     excel = win32com.client.Dispatch("Excel.Application")
    #     excel.Visible = True
    #     workbook = excel.Workbooks.Open(ruta_archivo)
    #     sheet = workbook.Sheets(1)

    #     def seleccionar_columna(nombre_columna):
    #         for cell in sheet.Rows(1):
    #             if cell.Value == nombre_columna:
    #                 sheet.Columns(cell.Column).Select()
    #                 break
                
    #     seleccionar_columna("cue_ncuenta")
    #     seleccionar_columna("rec_czona")
    #     seleccionar_columna("rec_tFechaProceso")
    #     seleccionar_columna("rec_tFechaRecepcion")
    #     seleccionar_columna("_puerto")
    #     # Conexion a PostgreSQL e insercion de datos

    #     # Calcular la fecha del di­a anterior

    #     # Cerrar el archivo de Excel y la aplicacion de Excel
    #     if 'workbook' in locals():
    #         workbook.Close(SaveChanges=False)
    #     if 'excel' in locals():
    #         excel.Quit()
    #     # Llamar a la funcion para insertar y mostrar los datos
    #     # consultarlistaAbonados(col_cue_ncuenta, tabla,col_rec_czona, col_rec_tFechaProceso, col_rec_tFechaRecepcion, col_puerto,nombre_proceso)


    # except Exception as e:
    #     logging.error("Error al procesar el archivo de Excel:", e)
    #     print("La inserción en la base de datos fue exitosa.")

def procesar_archivo_excel(descargas_path='~/Downloads', tabla=None, nombre_proceso=None, col1=None, col2=None, col3=None, col4=None, col5=None):
    try:
        #ir a la carpeta descargas
        print('dentro de la funcion: procesar_archivo_excel')
        
        descargas_path = os.path.expanduser(descargas_path)
        #Todos los archivos tipo excel en la carpeta descargas
        archivos_excel_en_descargas = glob.glob(os.path.join(descargas_path, '*.xlsx'))

        if archivos_excel_en_descargas:
            archivo_excel = max(archivos_excel_en_descargas, key=os.path.getctime)
            print(archivo_excel)
    except Exception as e:
        print('Ocurrio un error al encontrar la ruta del archivo excel: ', e)
    
    try:
        print('abrir archivo excel')
        # leer archivo excel
        excel = pandas.read_excel(archivo_excel)
        
        #seleccionar columna
        if nombre_proceso == 'baterias1' or 'baterias2':
            columna = excel[col1]
            for valor in columna:
                print(valor)
    except Exception as e:
        print('Ocurrio un error al abrir el archivo excel')


def recorrer_formulario_filtrar():
    try:
        pyautogui.press('tab')
        pyautogui.write(mes_y_ano, interval=0.1)
        time.sleep(1)
        pyautogui.press('tab')
        pyautogui.press('tab')
        for nombre_proceso, fecha_desde, fecha_hasta, hora_ejecucion, hora_desde, hora_hasta, codigo_alarma, col1, col2, col3, col4, col5, tabla in horarios_procesos:
        # PROCESo 2
            if  hora_ejecucion == hora_actual:
                logging.info(f'Proceso a ejecutar: {nombre_proceso}, Hora ejecucion: {hora_ejecucion}, Hora actual: {hora_actual}')
                print('nombre_proceso:')
                print('hora ejecucion:', hora_ejecucion)
                print('hora actual: ', hora_actual)            
                print('proceso: ', nombre_proceso, 'fecha_desde: ', fecha_desde, 'fecha_hasta: ', fecha_hasta, 'hora_ejecucion:', hora_ejecucion, 'hora_desde: ', hora_desde, 'hora_hasta:', hora_hasta, 'codigo_alarma: ', codigo_alarma)
                time.sleep(1)
                pyautogui.write(fecha_desde, interval=0.01)            
                time.sleep(1)
                pyautogui.press('tab')
                pyautogui.write(hora_desde, interval=0.01)
                time.sleep(1)
                
                pyautogui.press('tab')
                pyautogui.write(fecha_hasta, interval=0.01)
                time.sleep(1)
                
                pyautogui.press('tab')
                pyautogui.write(hora_hasta, interval=0.01)
                time.sleep(1)
                
                pyautogui.typewrite(['tab'] * 5)
                pyautogui.write(codigo_alarma, interval=0.01)
                time.sleep(1)
                pyautogui.press('down')
                pyautogui.press('enter')
                time.sleep(20)
                ruta_captura = obtener_captura_pantalla('filtrando_proceso.png', 'screenshots')
                mensaje_telegram('filtrando_proceso', ruta_captura, None, None, nombre_proceso, hora_ejecucion, None)
                break
            else:
                logging.error(f'El horario {hora_ejecucion} no coincide con la hora actual {hora_actual}')
                print(f'El horario {hora_ejecucion} no coincide con la hora actual {hora_actual}')
                mensaje_telegram('horarios_diferentes', None, None, None, nombre_proceso, hora_ejecucion, None)
        if ruta_captura:
            print('fuera del for, pero dentro del otro if')
            iniciar_filtro(tupla_postformulario[0])
            time.sleep(10)
            no_hay_eventos = obtener_ruta_imagenes('no_hay_eventos.png')
            coordenadas_no_hay_eventos = obtener_coordenadas_imagen_pantalla(no_hay_eventos)
            if coordenadas_no_hay_eventos != 'error':
                print(f'no hay eventos, nombre proceso: {nombre_proceso}')
                logging.error(f'no hay eventos, nombre proceso: {nombre_proceso}')
                ruta_captura = obtener_captura_pantalla('no_hay_eventos.png', 'screenshots')
                mensaje_telegram('no hay eventos', ruta_captura, None, None, nombre_proceso, None, None)
                logging.error(f'No se encontraron eventos para descargar el archivo excel en el proceso de {nombre_proceso}, mensaje enviado al telegram')
                print(f'No se encontraron eventos para descargar el archivo excel en el proceso de {nombre_proceso}, mensaje enviado al telegram')
                logging.error('Se cancela la ejecucion del proceso censel')
                os.system("taskkill /f /im chrome.exe")
            else:
                print('ejecuntando tupla_postforulario')
                for img in tupla_postformulario:
                    iniciar_filtro(img)
                    if img == 'exportar_a_csv.png':
                        time.sleep(2)
                        os.system("taskkill /f /im chrome.exe")
                        print('Se finaliza el proceso en la pagina de censel y se cierra el navegador.')
                        mensaje_telegram('cerrar_chrome', None, None, None, None, None ,None)
                        print('nombre_proceso: ', nombre_proceso)
                        print('tabla: ', tabla)
                        procesar_archivo_excel(tabla=tabla, nombre_proceso=nombre_proceso, col1=col1, col2=col2, col3=col3, col4=col4, col5=col5)
                        print('antes del exit.')
                        print('tabla: ', tabla)
                        print('nombre_proceso: ', nombre_proceso)
                        # exit()
    except Exception as e:
        logging.error(f'Error en la funcion: recorrer_formulario_filtrar: {e}')



# def validar_descarga_archivo()
#             print('hola3')
#             iniciar_filtro(tupla_postformulario[0])
#             time.sleep(10)
#             no_hay_eventos = obtener_ruta_imagenes('no_hay_eventos.png')
#             coordenadas_no_hay_eventos = obtener_coordenadas_imagen_pantalla(no_hay_eventos)
#             if coordenadas_no_hay_eventos != 'error':
#                 print(f'no hay eventos, nombre proceso: {nombre_proceso}')
#                 logging.error(f'no hay eventos, nombre proceso: {nombre_proceso}')
#                 ruta_captura = obtener_captura_pantalla('no_hay_eventos.png', 'screenshots')
#                 mensaje_telegram('no hay eventos', ruta_captura, None, None, nombre_proceso, None, None)
#                 logging.error(f'No se encontraron eventos para descargar el archivo excel en el proceso de {nombre_proceso}, mensaje enviado al telegram')
#                 print(f'No se encontraron eventos para descargar el archivo excel en el proceso de {nombre_proceso}, mensaje enviado al telegram')
#                 logging.error('Se cancela la ejecucion del proceso censel')
#                 os.system("taskkill /f /im chrome.exe")
#             else:
#                 print('ejecuntando tupla_postforulario')
#                 for img in tupla_postformulario:
#                     iniciar_filtro(img)
#                     if img == 'exportar_a_csv.png':
#                         time.sleep(20)
#                         if os.system("taskkill /f /im chrome.exe"):
#                             mensaje_telegram('cerrar_chrome', None, None, None, None, None ,None)
#                             #procesar_archivo_excel(r'C:\Users\auxsenadesarrollo\Downloads\reportehistoricohtml.xlsx', 'replica_registro_codigos_seguridad', nombre_proceso)
#                             exit()
#             # return




def iniciar_sesion():
    try:
        # ABRIR NAVEGADOR
        logging.info (f"SE INICIA PROCESO A LAS {hora_actual}")
        mensaje_telegram('inicio', None, None, None, None, None, None)
        pyautogui.hotkey('win', 'r')
        time.sleep(2)
        pyautogui.write('chrome --start-maximized', interval=0.01)
        time.sleep(2)
        pyautogui.press('enter')
        time.sleep(2)
        print('Hola mundo')
        
        # ABRIR CENSEL  
        time.sleep(2)
        pyautogui.write("censelc.ultrasecuritysolution.com", interval=0.01)
        time.sleep(1)
        pyautogui.press('enter')
        time.sleep(2)

        # si habia una sesion abierta
        for img in tupla_inicar_sesion:
            iniciar_filtro(img)
        
            # INICIAR SESION
        time.sleep(2)
        pyperclip.copy('@')
        #pyautogui.hotkey('tab')
        pyautogui.write('oscartorres', interval=0.01)
        time.sleep(2)
        # pegar el @
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(2)
        pyautogui.write('consuerte.com', interval=0.01)
        pyautogui.hotkey('tab')
        time.sleep(2)
        pyautogui.write('Oscar13579', interval=0.01)
        time.sleep(2)
        pyautogui.hotkey('tab')
        pyautogui.press('enter')

        time.sleep(10)
        for img in tupla_filtro:
            logging.info('Se inicia ingreso a reportes web')
            iniciar_filtro(img)
        
        recorrer_formulario_filtrar() 
    except Exception as error: 
        logging.error(' No se pudo inicar sesion: ')


#procesar_archivo_excel(r'C:\Users\auxsenadesarrollo\Downloads\reportehistoricohtml.xlsx','replica_registro_codigos_seguridad', 'baterias')
# replica_seg_control_novedades
iniciar_sesion()
# procesar_archivo_excel()